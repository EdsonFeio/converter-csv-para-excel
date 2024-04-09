import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle

def converter_csv_para_excel(pasta_csv, pasta_excel):
    # Verifica se a pasta de destino existe, se não, cria-a
    if not os.path.exists(pasta_excel):
        os.makedirs(pasta_excel)

    sucesso = True  # Variável para verificar se a conversão foi bem-sucedida para todos os arquivos

    # Itera sobre todos os arquivos na pasta CSV
    for arquivo_csv in os.listdir(pasta_csv):
        if arquivo_csv.endswith('.csv'):
            # Caminho completo para o arquivo CSV
            caminho_csv = os.path.join(pasta_csv, arquivo_csv)

            # Lê o arquivo CSV usando pandas
            dados = pd.read_csv(caminho_csv)

            # Cria um novo arquivo Excel usando openpyxl
            wb = Workbook()
            ws = wb.active

            # Define estilo para células numéricas
            number_style = NamedStyle(name='number')
            number_style.number_format = '0'  # Define o formato como número

            # Adiciona os dados ao arquivo Excel
            for row in dataframe_to_rows(dados, index=False, header=True):
                ws.append(row)

            # Define o estilo para as células numéricas
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.style = number_style

            # Ajusta a largura das colunas ao tamanho do maior conteúdo da coluna
            for column_cells in ws.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length + 2

            # Ajusta a altura das linhas para acomodar o conteúdo
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        ws.row_dimensions[cell.row].height = (len(str(cell.value)) // 50 + 1) * 12

            # Salva o arquivo Excel na pasta de destino
            nome_arquivo_excel = os.path.splitext(arquivo_csv)[0] + '.xlsx'
            caminho_excel = os.path.join(pasta_excel, nome_arquivo_excel)
            wb.save(caminho_excel)

            print(f"Arquivo {arquivo_csv} convertido com sucesso.")

    if sucesso:
        print("Todos os arquivos foram convertidos com sucesso.")
    else:
        print("Falha ao converter algum arquivo.")

# Caminho para a pasta contendo os arquivos CSV
caminho_pasta_csv = r'C:\Users\edson.feio\Downloads\pasta_com_csv'

# Caminho para a pasta onde os arquivos Excel de saída serão salvos
caminho_pasta_excel = r'C:\Users\edson.feio\Downloads\pasta_com_excel'

# Converte os arquivos CSV para Excel
converter_csv_para_excel(caminho_pasta_csv, caminho_pasta_excel)
